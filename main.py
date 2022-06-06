#job Scrapper MA1
'''https://t.me/job_scrapper_ma1_bot'''

import logging
# from telegram import Update
#pip install python-telegram-bot
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters,CallbackContext
from time import sleep, time
import json
from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import *

TOKEN = open("token/token.txt", "r").read()

#let's add these
# - put the token in a seperate file
# - use threading used in project excel/whats
# - asking questions like skills and posting time
# - auto correct the words using a library
# - match the jobs with the skills to get better results
# send a statistics message with the excel file ex: num of jobs


# Enable logging
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    level=logging.INFO)
logger = logging.getLogger(__name__)


def start(update, context):
    """Send a message when the command /start is issued."""
    update.message.reply_text(f"Hello {update['message']['chat']['first_name']} {update['message']['chat']['last_name']}!!")
    update.message.reply_text(f'Hello in the Job scrapper bot! Write j then the job you need to get you an excel of 2000 job for you\nExample: j python engineer')

def help(update, context):
    """Send a message when the command /help is issued."""
    update.message.reply_text("Hello in the Job scrapper bot! Write j then the job you need to get you an excel of 2000 job for you\nExample: j python engineer")


def echo(update, context):

    """Echo the user message."""
    name = f"{update['message']['chat']['first_name']} {update['message']['chat']['last_name']}"
    # saveInExcel(update)

    for i in range(1):
        update.message.reply_text(update.message.text,  reply_to_message_id=update.message.message_id)

    if update.message.chat.id != 1074890834:
        print(00)
        theMsg = f"{name} - @{update.message.chat.username} :\n{update.message.text}"
        print(1)
        context.bot.send_message('1074890834',theMsg)
        print(11)

    if update.message.text.split()[0].lower() in 'j':
        job=' '.join(update.message.text.split()[1:])
        update.message.reply_text(f'we are finding jobs for {job}')

        x=FindJob(job,name)

        update.message.reply_text(f'Here are 2000 jobs in EXCEL:\n{x}\nGenerated by MA-bot')
        context.bot.sendDocument(update['message']['chat']['id'], document=open(x, 'rb'))

    else:
        update.message.reply_text(f"Hello {name}!\nPlease write j in the beginning")

def error(update, context):
    """Log Errors caused by Updates."""
    logger.warning('Update "%s" caused error "%s"', update, context.error)
    logger.warning(context.error)

def download(update, context):
    update.message.reply_text('haha')
    context.bot.sendDocument(update['message']['chat']['id'],document=open('s.xlsx','rb'))


def FindJob(name_,user):
    wb = Workbook()
    ws = wb.active
    key='+'.join(name_.split())
    c=3
    for i in range(1,2):

        html= requests.get(f'https://www.timesjobs.com/candidate/job-search.html?from=submit&actualTxtKeywords={key}&searchBy=0&rdoOperator=OR&searchType=personalizedSearch&luceneResultSize={100}&postWeek=60&txtKeywords=python&pDate=I&sequence={i}').text

        soup=BeautifulSoup(html,'lxml')


        jobs=soup.find_all('li', class_= "clearfix job-bx wht-shd-bx")

        for job in jobs:
            c+=1

            try:name = job.find('a').text.strip()
            except:name='x'

            try:years=job.find('ul',class_="top-jd-dtl clearfix").find('li').text.split('card_travel')[1]
            except:years='x'

            try:location = job.find('ul',class_="top-jd-dtl clearfix").find_all('li')[1].find('span').text
            except:
                try: location = job.find('ul',class_="top-jd-dtl clearfix").find_all('li')[2].find('span').text
                except:location='x'

            try:cmp=job.find('h3',class_="joblist-comp-name").text.strip().split('\n')[0].strip()
            except:cmp='x'

            try:skills=job.find('span',class_="srp-skills").text.strip().replace(' ','').replace(',',' & ')
            except:skills='x'

            try:posted=job.find('span',class_="sim-posted").text.strip().replace('Posted ','')
            except:posted='x'

            try:describtion=job.find('ul',class_="list-job-dtl clearfix").find('li').text.strip().replace('Job Description:','').replace('More Details','').strip()
            except:describtion='x'

            try:link=job.find('a')['href']
            except:link='x'

            print(f'''job {c-3}: ({name_} - {user}.xlsx)
        {name}
        {cmp}
        {location}
        {years}
        {posted}
        {describtion}
        {skills}
        {link}
        ''')
            ws[f'A{c}'] = name
            ws[f'B{c}'] = cmp
            ws[f'C{c}'] =location
            ws[f'D{c}'] = years
            ws[f'E{c}'] =posted
            ws[f'F{c}'] =describtion
            ws[f'G{c}'] =skills
            ws[f'H{c}'].hyperlink=link
            ws[f'H{c}'].font = Font(color='000000FF', underline="single")
    ws['A3']= 'NAME'
    ws['B3']= 'COMPANY'
    ws['C3']= 'LOCATION'
    ws['D3']= 'EXPERIENCE TIME'
    ws['E3']= 'POSTED WHEN'
    ws['F3']= 'DESCRIPTION'
    ws['G3']= 'SKILLS REQUIRED'
    ws['H3']= 'LINK'
    ws['A1'] = f'This excel is generated by Ashmawy bot to scrap jobs data from "Times Jobs" about {name_} and is asked by {user}'


    ws.merge_cells('A1:H2')
    # ws.column_dimensions['A:H'].width=20
    for c in 'ABCDEFGH':
        ws.column_dimensions[f'{c}'].width=20



    file_name=f"{name_} - {user}.xlsx"
    wb.save("output files/"+file_name)
    return file_name

def saveInExcel(update):
    workbook = load_workbook('Msg-History.xlsx')
    sheet = workbook.active
    last = sheet.max_row + 1

    # js = json.loads(str(update))

    sheet[f'A{last}'] = str(update.message.date)
    sheet[f'B{last}'] = update.message.chat.username
    sheet[f'C{last}'] = str(update.message.chat.first_name) + " " + str(update.message.chat.last_name)
    sheet[f'D{last}'] = update.message.chat.id
    sheet[f'E{last}'] = update.message.text
    sheet[f'F{last}'] = update.message.message_id
    sheet[f'G{last}'] = str(update)
    workbook.save('Msg-History.xlsx')


def main():
    updater = Updater(TOKEN, use_context=True)
    dp = updater.dispatcher
    dp.add_handler(CommandHandler("start", start))
    dp.add_handler(CommandHandler("help", help))
    dp.add_handler(CommandHandler("d", download))
    try:
        dp.add_handler(MessageHandler(Filters.text, echo))
    except:
        dp.add_handler(MessageHandler(Filters.text, echo))
    dp.add_error_handler(error)
    updater.start_polling()
    updater.idle()


if __name__ == '__main__':
    main()
